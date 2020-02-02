# coding=utf-8
from openpyxl import load_workbook, workbook


class CreateExcel:
    def __init__(self, work_book, sheet):
        """
        :param work_book: 工作簿
        :param sheet: 表单
        """
        self.book_name = work_book
        self.sheet = sheet

    def create_excel(self):
        wb = workbook.Workbook()
        wb.create_sheet(self.sheet)
        wb.save(self.book_name)


class DoExcel:
    def __init__(self, work_book, sheet):
        self.work_book = work_book
        self.sheet = sheet

    """列表嵌套列表"""

    def read_excel_list(self):
        wb = load_workbook(self.work_book)
        sheet = wb[self.sheet]
        case_list = list()
        for i in range(1, sheet.max_row):
            case = list()
            for j in range(1, sheet.max_column):
                value = sheet.cell(i, j).value
                case.append(value)
            case_list.append(case)
        return case_list

    """列表嵌套字典"""

    def read_excel_dict(self):
        wb = load_workbook(self.work_book)
        sheet = wb[self.sheet]
        rows = sheet.rows
        print(rows)
        first_line = list(map(lambda x: x.value, next(rows)))

        case_list = list()
        for i in range(1, sheet.max_row):
            """每一行的值"""
            lines = list(map(lambda y: y.value, next(rows)))
            """聚合首行和循环得到的每一行"""
            data = list(zip(first_line, lines))
            case_list.append(dict(data))
        print(case_list)

    def write_excel(self, row, column, value):
        wb = load_workbook(self.work_book)
        sheet = wb[self.sheet]
        sheet.cell(row=row, column=column).value = value
        wb.save(self.work_book)


if __name__ == '__main__':
    de = DoExcel("testcase.xlsx", "login")
    de.read_excel_dict()


    # res = de.read_excel()
    # de.write_excel(3, 1, 22)
    # res = de.read_excel_list()
    # print(res)

